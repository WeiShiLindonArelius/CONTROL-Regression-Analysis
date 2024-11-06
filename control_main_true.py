import pandas as pd
from Teams import Team, generate_lineups_six_to_four
from Players import PlayerSeason
from contests import round_robin
from colorama import Fore, Style
from dump_pickle import dump_pkl
from load_pickle import season_wipe, load_pkl
import time
from random import choice, shuffle
from leagues import league_season, user_draft, player_changes, grade_players, double_elim_8, double_elim_16, swiss_format, write_champ
from Games import best_of, enablePrint
import concurrent.futures
from stat_functions import season_stats, best_of_stats, region_mvp, QUERY, initiate_databases, finalize_series_data
from statistics import mean
from collections import OrderedDict
import re
from openpyxl import load_workbook
import sqlite3

#main begins on line 387, and the first season begins on line 468
#to change manual/auto team sorting, line 480 change manual=
#line 1200 contains variable to toggle rosters being written
#todo Slashers are not being drafted, either because they're actually bad or because xWAR is inaccurate

#NO_SQL is in Players.py

#done make it so that teams who miss the playoffs in the region have the option to replace two players
# if there is a better option left in the draft pool
# after the first draft, all teams who missed the regional playoffs and teams who made it to the
# pre-qualifying tournament or further the previous season will have a second draft pick
# if there is a player in the draft list better than one of their players (based on xWAR) they draft them
# teams from all regions select from one massive draft class called secondary
# teams that miss the playoffs will be put into one pool and shuffled, and teams who made PQ or further will be
# shuffled in another pool. the missed_playoffs pool will select first.

#teams will have a trait called second_pick naturally set to False but set to true if they miss the regional playoffs
#or make it to the pre-qualifier or universal qualifier
#after all the first drafts are over, I will run second_draft() with a list of all teams who have second_pick == True

#idea: regulate player statistics to "per 50" statistics which will stay proportional to lower and higher game lengths
#a season with a game length of 25 ticks should have all stats (kills pg, damage pg, deaths pg, etc.) multiplied by 2
#Every season object must have a value for the length (in ticks) of the games played in this season
#Player seasons are generated from player values, so every player must have a game_length stat
#because the game length stat originates from game(), I can make it so that every player who plays a game has their game_length
#stat set to the length of the game. this will result in redundant code, so I can add a check for this

#GLOBAL VARIABLES
avg_stats_df = pd.DataFrame(columns=['Kills', 'Deaths', 'Damage', 'Effect', 'Overkill', 'Mitigated'])


def weighted_averages(team):  # takes in a team and calculates weighted average of each stat, returns a dataframe
    #despite being called weighted_averages, this produces a full dataframe for a team season
    #todo add a function to this which extracts part of a team's history that shows how their season ended
    # could potentially add sub-function which quantifies how good a season is based on its finish, which could be
    # useful for tracing team stats to success

    final_dict = {'Team': team.name, 'Power': None, 'Attack Damage': None, 'Attack Speed': None, 'Critical %': None,
                  'Critical X': None, 'Health': None, 'Spawn Time': 0, 'Lineup Wins': team.wins,
                  'Lineup Losses': team.losses
        , 'Match Wins': team.match_wins, 'Match Losses': team.match_losses, 'Match Draws': team.match_draws}

    power_list = [player.power for player in team.players]
    total_power = (power_list[0] * 9) + (power_list[1] * 7) + (power_list[2] * 6) + (power_list[3] * 6) + (
                power_list[4] * 4) + (power_list[5] * 4)
    final_dict['Power'] = round((total_power / 36), 2)

    atk_dmg_list = [player.atk_dmg for player in team.players]
    total_atk_dmg = (atk_dmg_list[0] * 9) + (atk_dmg_list[1] * 7) + (atk_dmg_list[2] * 6) + (atk_dmg_list[3] * 6) + (
                atk_dmg_list[4] * 4) + (atk_dmg_list[5] * 4)
    final_dict['Attack Damage'] = round((total_atk_dmg / 36), 2)

    atk_spd_list = [player.atk_spd for player in team.players]
    total_atk_spd = (atk_spd_list[0] * 9) + (atk_spd_list[1] * 7) + (atk_spd_list[2] * 6) + (atk_spd_list[3] * 6) + (
                atk_spd_list[4] * 4) + (atk_spd_list[5] * 4)
    final_dict['Attack Speed'] = round((total_atk_spd / 36), 2)

    crit_pct_list = [player.crit_pct for player in team.players]
    total_crit_pct = (crit_pct_list[0] * 9) + (crit_pct_list[1] * 7) + (crit_pct_list[2] * 6) + (
                crit_pct_list[3] * 6) + (crit_pct_list[4] * 4) + (crit_pct_list[5] * 4)
    final_dict['Critical %'] = round((total_crit_pct / 36), 5)

    crit_x_list = [player.crit_x for player in team.players]
    total_crit_x = (crit_x_list[0] * 9) + (crit_x_list[1] * 7) + (crit_x_list[2] * 6) + (crit_x_list[3] * 6) + (
                crit_x_list[4] * 4) + (crit_x_list[5] * 4)
    final_dict['Critical X'] = round((total_crit_x / 36), 3)

    health_list = [player.max_health for player in team.players]
    total_health = (health_list[0] * 9) + (health_list[1] * 7) + (health_list[2] * 6) + (health_list[3] * 6) + (
                health_list[4] * 4) + (health_list[5] * 4)
    final_dict['Health'] = round((total_health / 36), 3)

    spawn_list = [player.spawn_time for player in team.players]
    total_spawn = (spawn_list[0] * 9) + (spawn_list[1] * 7) + (spawn_list[2] * 6) + (spawn_list[3] * 6) + (
                spawn_list[4] * 4) + (spawn_list[5] * 4)
    final_dict['Spawn Time'] = round((total_spawn / 36), 2)

    return pd.DataFrame([final_dict])

def team_season_dataframe(teams, season_no):
    #writes the average team player innate stats and their season record to a file
    #despite the name, this takes in a list of teams and writes them out after making a DF

    team_stats_df = pd.DataFrame(columns=['Team', 'Power', 'Attack Damage', 'Attack Speed', 'Critical %', 'Critical X',
                                              'Health', 'Spawn Time', 'Lineup Wins', 'Lineup Losses',
                                              'Match Wins', 'Match Losses', 'Match Draws'])
    for team in teams:
        team_stats_df = pd.concat([team_stats_df, weighted_averages(team)])

    path = "ControlAverageStats.xlsx"

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        team_stats_df.to_excel(writer, sheet_name='TeamWeightedAverages', index=False)

def sort_players_by_xWAR(team):
    team.players.sort(key=lambda pl: pl.xWAR, reverse=True)

def sort_all_players(leagues, manual):
    #teams have a boolean called mine
    def move_player(old, new, swap_team):
        swap_team.players[old], swap_team.players[new] = swap_team.players[new], swap_team.players[old]


    for league in leagues:
        for team in league:
            grade_players(team.players, is_team=True)
            if team.mine:
                enablePrint()
                if manual:
                    print(f"Here is the current order of your players ({team.name}):" + Fore.RESET)
                    i = 0
                    for player in team.players:
                        print(f"Slot {i}, {player}")
                        i += 1
                    while True:
                        user_choice = input(Fore.BLUE + "Would you like to move anything? If not, press N.\n"
                                                        "To move a player, press their Slot Number.")
                        if user_choice == 'N' or user_choice == 'n':
                            break
                        else:
                            user_choice = int(user_choice)
                        old_slot = user_choice
                        new_slot = int(input("What will their new slot be?"))
                        move_player(old_slot, new_slot, team)
                        print(Fore.RESET + "Here is the new order of your players:")
                        i = 0
                        for player in team.players:
                            print(f"Slot {i}, {player}")
                            i += 1
            team.lineups = generate_lineups_six_to_four(team.players)



def write_to_file(filename=None, words=None, mode='w', error=False):
    if error:
        filename = 'error_output'
        mode = 'a'

    with open(filename, mode) as f:
        f.write(words + '\n')


def clean_file(file_path):
    with open(file_path, 'r+') as file:
        lines = file.readlines()
        file.seek(0)
        file.truncate()

        delete_next_line = False

        for line in lines:
            # Check if the line contains a colon
            if ':' in line:
                delete_next_line = True
                continue

            # Check if the line should be deleted based on the previous line
            if delete_next_line:
                delete_next_line = False
                continue

            # Check if the line is empty
            if line.strip() == '':
                continue

            # Write the line back to the file
            file.write(line)

def format_champion_line(line):
    # Define a regex pattern to match the line format
    if "Regional" in line:

        pattern = re.compile(r'^S(\d+)\s(.+?)\sRegional\sChampion:\s(.+)$')

        # Use the regex pattern to match and extract groups
        match = pattern.match(line)

        # If the line matches the pattern, format it as required
        if match:
            seed = match.group(1)
            region = match.group(2)
            champion = match.group(3)
            formatted_line = f"{champion} (S{seed} {region} Regional Champion)"
            return formatted_line
        else:
            # If the line doesn't match the expected format, return the original line
            return line.strip()
    else:
        # Define a regex pattern to match the required line format
        pattern = re.compile(r'^S(\d+)\s(.+?)\sChampion:\s(.+)$')

        # Use the regex pattern to match and extract groups
        match = pattern.match(line)

        # If the input string matches the format, format it as required
        if match:
            seed = match.group(1)
            champion = match.group(3)
            formatted_string = f"{champion} (S{seed} {champion} Champion)"
            return formatted_string
        else:
            # If the input string doesn't match the expected format, return the original string
            return line.strip()

#todo urgent: for each league, kills and deaths should be equal. for best_stats, this is not the case. after running each season and testing this, it seems it is only happening to one league
#where kills and deaths are unequal. my task now is to find where it is happening and why
# grade_seasons is called in two different contexts: every time league_season() runs, grade_seasons runs
# and every season, AFTER the universal league, grade_seasons runs via best_of_stats
# my hypothesis is that the problem lies either with best_of_stats, or the nature of the full season list
#which is created by season_stats
#todo IMPORTANT: through debugging I have found that the differential between kills and deaths occurs DURING GRADE_SEASONS and ONLY when it is run via best_of_stats()
# by testing the kills and deaths values for seasons upon their entry into the large stats_list through season_stats(), I found that death numbers are significantly greater
# IN TOTAL before grade_seasons() (and obviously best_of_stats()) is ever run. so the problem occurs before this point, either in the creation or the handling of season data
# by testing season_stats(), I discovered that the transfer of statistics from player to season objects is flawless and not part of the issue
# this also gave me a list of every single player and their kills/deaths. I expected to find some anomalies, but to no avail
# I can conclude that somehow, when you add all of the player seasons together (and the definition of season might be an issue) there are more deaths than kills
# REVELATION: the issue is almost certainly due somehow to season_wipe, but I won't go any further for now

def sort_champions_by_seed(file_path='champs'):
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Create a list to store tuples (seed, champion_info)
    champions = []

    index = 0
    for line in lines:

        if line.startswith('Entered playoffs as'):
            parts = line.split()
            seed = int(parts[-2])
            champion_info = lines[index-1].strip()
            champion_info = format_champion_line(champion_info)
            champions.append((seed, champion_info))

        index += 1

    # Sort the champions list in descending order based on seed
    champions.sort(reverse=True, key=lambda x: x[0])

    # Print the sorted champions
    clean_file(file_path)
    with open(file_path, 'a') as file:
        file.write('\n')
        for champion in champions:
            file.write(f"{champion[1]} Entered playoffs as {champion[0]} seed.\n")
        file.write('\n\n')


def remove_duplicates_ordered(upl_standings):
    seen_teams = OrderedDict()
    unique_standings = []

    for team in upl_standings:
        if team not in seen_teams:
            unique_standings.append(team)
            seen_teams[team] = None

    return unique_standings


def extract_high_seed(text):
    int_values = [int(match.group()) for match in re.finditer(r'\b\d+\b', text)][:2]
    high_seed = min(int_values)
    return high_seed

def get_upsets(upset_list, upset_count):
    #upset[0]: a string of the upset ex: "12(Universal_Draconians) def. 6(Web-of-Nations_Oni) by a score of 547-504 (R2L Universal Playoffs)"
    #upset[1]: integer, difference between seeds
    #upset[2]: integer, highest seed in the match (the seed which was upset)


    clear_file('upsets')
    seed_diff_list = []
    non_rel_upsets = []
    uni_upsets = []
    high_seed_dict = {}

    for upset in upset_list:
        high_seed_dict[upset] = extract_high_seed(upset[0])
        seed_diff_list.append(upset[1])
        if "Relegation" not in upset[0]:
            non_rel_upsets.append(upset)
            if any(term in upset[0] for term in ["Universal Round of 16", "Universal Quarterfinals", "Universal Semifinals", "Universal Finals"]):
                uni_upsets.append(upset)
    seed_diff_avg = mean(seed_diff_list)

    with open('upsets', 'a') as u:
        u.write(f"So far, there have been {len(upset_list)} upsets, with an average seed differential of {seed_diff_avg:.3f}.\n\n")
        upset_list.sort(key=lambda l: (l[1], -(high_seed_dict[l])), reverse=True)
        non_rel_upsets.sort(key=lambda l: (l[1], -(high_seed_dict[l])), reverse=True)
        uni_upsets.sort(key=lambda l: (l[1], -(high_seed_dict[l])), reverse=True)

        if len(uni_upsets) >= 5:
            u.write("\n5 Largest Universal League Upsets\n")
            for i in range(5):
                u.write(f"{uni_upsets[i][0]}\n")
        else:
            u.write(f"\n{len(uni_upsets)} Largest Universal League Upsets\n")
            for i in range(len(uni_upsets)):
                u.write(f"{uni_upsets[i][0]}\n")

        if len(non_rel_upsets) >= 5:
            u.write("\n5 Largest Upsets (not incl. Universal Relegation)\n")
            for i in range(5):
                u.write(f"{non_rel_upsets[i][0]}\n")
        else:
            u.write(f"\n{len(non_rel_upsets)} Largest Upsets (not incl. Universal Relegation)\n")
            for i in range(len(non_rel_upsets)):
                u.write(f"{non_rel_upsets[i][0]}\n")
        if len(upset_list) >= 3:
            u.write("\n3 Largest Overall Upsets\n")
            for i in range(3):
                u.write(f"{upset_list[i][0]}\n")
        else:
            u.write(f"\n{len(upset_list)} Largest Overall Upsets\n")
            for i in range(len(upset_list)):
                u.write(f"{upset_list[i][0]}\n")

        u.write("\nAll Upsets:\n")
        for t in upset_list:
            u.write(f"{t[0]}\n")



GENERAL_OUTPUT = False

start = time.time()

# try def __print__(): to get rid of colorama


def clear_file(filename, excel=False):
    if not excel:
        with open(filename, 'w') as c:
            c.write('')
    else:
        workbook = load_workbook(filename)

        # Loop through every sheet in the workbook
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]

            # Loop through every row in the sheet
            for row in worksheet.iter_rows():
                # Loop through every cell in the row
                for cell in row:
                    # Clear the cell value
                    cell.value = None

        # Save the cleared workbook
        workbook.save(filename)

def create_teams(count,region='None',season_count=0):
    TEAMS = []
    for i in range(count):
        temp = Team(region,season_count=season_count)
        TEAMS.append(temp)
    return TEAMS

def ordinal_string(n: int) -> str:
    if 10 <= n % 100 < 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

def main():

    # file clearing:
    clear_file('my_teams')
    clear_file('region_mvp.txt')
    clear_file('champs')
    clear_file('error_output')
    clear_file('upsets')
    clear_file('draft_list')
    clear_file('draft_history')
    clear_file('player_trait_data')

    clear_file('PlayerSeasons.xlsx', excel=True)
    clear_file("ControlAverageStats.xlsx", excel=True)

    my_team_count = 4

    SEASONS = 30

    season_stats_list = {}
    champ_list = []


    all_teams = []
    #this is the only list which should contain every single season list, from every single season, period

    uni_stats_list = {}
    dw_stats_list = {}
    sc_stats_list = {}
    ds_stats_list = {}
    wof_stats_list = {}
    iw_stats_list = {}
    cl_stats_list = {}
    hc_stats_list = {}
    sh_stats_list = {}

    upset_list = []
    upset_count = 0

    for i in range(SEASONS+1):
        uni_stats_list[i] = list()
        dw_stats_list[i] = list()
        sc_stats_list[i] = list()
        ds_stats_list[i] = list()
        wof_stats_list[i] = list()
        iw_stats_list[i] = list()
        cl_stats_list[i] = list()
        hc_stats_list[i] = list()
        sh_stats_list[i] = list()

    use_saved = False
    #this should work as of 11/02/2024

    if use_saved:
        try:
            upl_standings, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams, season_count = load_pkl()
            uni_teams = upl_standings
        except ValueError:
            use_saved = False
            print(Fore.RED + "Failed to save values, starting new iteration.")
    if not use_saved:
        season_count = 0
        #23 to account for 3 trait-based teams
        uni_teams = create_teams(23, "Universal", season_count=season_count)
        cosmic_team = Team("Cosmic", mine=False, pre_name="Slashers",season_count=0)
        undead_team = Team("Tartarus", mine=False, pre_name="Undead",season_count=0)
        reflect_team = Team("Beyond", mine=True, pre_name="Reflectors",season_count=0)
        uni_teams.append(cosmic_team)
        uni_teams.append(undead_team)
        uni_teams.append(reflect_team)
        for team in uni_teams:
            team.history[season_count] = ""
        upl_standings = league_season(uni_teams, use_saved=False, season_count=0,upset_list=upset_list,upset_count=upset_count, region="Universal")
        dw_teams = create_teams(20, "Darkwing", season_count=season_count)
        sc_teams = create_teams(20, "Shining-Core", season_count=season_count)
        ds_teams = create_teams(20, "Diamond-Sea", season_count=season_count)
        wof_teams = create_teams(20, "Web-of-Nations", season_count=season_count)
        iw_teams = create_teams(20, "Ice-Wall", season_count=season_count)
        cl_teams = create_teams(20, "Candyland", season_count=season_count)
        hc_teams = create_teams(20, "Hell's-Circle", season_count=season_count)
        sh_teams = create_teams(20, "Steel-Heart", season_count=season_count)
        for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
            lab_teams = create_teams(2,'Labyrinth', season_count=season_count)
            league.append(lab_teams[0])
            league.append(lab_teams[1])
        for _ in range(my_team_count):
            chance = choice([dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams])
            my_team = choice(chance)
            my_team.make_mine(my_team.name)
            print('\n' + Fore.BLUE + f"{my_team.name} is your team.")
        print(Fore.RESET)

    def regional_leagues(dw_teams,sc_teams,ds_teams,wof_teams,iw_teams,cl_teams,hc_teams,sh_teams,season_count,champ_list):
            pre_qualif_tournament = []
            last_stand_tournament = {'6 Seeds' : [], '7 Seeds' : [], '8 Seeds' : []}

            print(Fore.GREEN + "DARKWING REGION, " + Fore.RESET, end='')
            dw_teams = league_season(dw_teams, False, season_count=season_count, final_reversed=False,
                                     region='Darkwing Regional', stats_list=dw_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            dw_qualified = dw_teams[:8]


            print(Fore.GREEN + "SHINING CORE REGION, " + Fore.RESET, end='')
            sc_teams = league_season(sc_teams, False, season_count=season_count, final_reversed=False,
                                     region='Shining-Core Regional', stats_list=sc_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            sc_qualified = sc_teams[:8]

            print(Fore.GREEN + "DIAMOND SEA REGION, " + Fore.RESET, end='')
            ds_teams = league_season(ds_teams, False, season_count=season_count, final_reversed=False,
                                     region='Diamond-Sea Regional', stats_list=ds_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            ds_qualified = ds_teams[:8]

            print(Fore.GREEN + "WEB OF NATIONS, " + Fore.RESET, end='')
            wof_teams = league_season(wof_teams, False, season_count=season_count, final_reversed=False,
                                      region='Web-of-Nations Regional', stats_list=wof_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            wof_qualified = wof_teams[:8]

            print(Fore.GREEN + "ICE WALL REGION, " + Fore.RESET, end='')
            iw_teams = league_season(iw_teams, False, season_count=season_count, final_reversed=False,
                                     region='Ice-Wall Regional', stats_list=iw_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            iw_qualified = iw_teams[:8]

            print(Fore.GREEN + "CANDYLAND REGION, " + Fore.RESET, end='')
            cl_teams = league_season(cl_teams, False, season_count=season_count, final_reversed=False,
                                     region='Candyland Regional', stats_list=cl_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            cl_qualified = cl_teams[:8]

            print(Fore.GREEN + "HELL'S CIRCLE, " + Fore.RESET, end='')
            hc_teams = league_season(hc_teams, False, season_count=season_count, final_reversed=False,
                                     region="Hell's-Circle Regional", stats_list=hc_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            hc_qualified = hc_teams[:8]

            print(Fore.GREEN + "STEEL HEART REGION, " + Fore.RESET, end='')
            sh_teams = league_season(sh_teams, False, season_count=season_count, final_reversed=False,
                                     region="Steel-Heart Regional", stats_list=sh_stats_list,upset_list=upset_list,upset_count=upset_count, champ_list=champ_list)
            sh_qualified = sh_teams[:8]

            dw_champ = dw_qualified[0]
            regional_champs.append(dw_champ)
            sc_champ = sc_qualified[0]
            regional_champs.append(sc_champ)
            ds_champ = ds_qualified[0]
            regional_champs.append(ds_champ)
            wof_champ = wof_qualified[0]
            regional_champs.append(wof_champ)
            iw_champ = iw_qualified[0]
            regional_champs.append(iw_champ)
            cl_champ = cl_qualified[0]
            regional_champs.append(cl_champ)
            hc_champ = hc_qualified[0]
            regional_champs.append(hc_champ)
            sh_champ = sh_qualified[0]
            regional_champs.append(sh_champ)

            # trans_region = ['Darkwing', 'Shining-Core', 'Diamond-Sea', 'Web-of-Nations', 'Ice-Wall', 'Candyland', "Hell's-Circle", 'Steel-Heart']
            count = 0
            for region in [dw_qualified, sc_qualified, ds_qualified, wof_qualified, iw_qualified, cl_qualified,
                           hc_qualified, sh_qualified]:
                for number in [1, 2, 3, 4]:
                    pre_qualif_tournament.append(region[number])

                last_stand_tournament['6 Seeds'].append(region[5])
                last_stand_tournament['7 Seeds'].append(region[6])
                last_stand_tournament['8 Seeds'].append(region[7])
                count += 1

            for region in [dw_qualified, sc_qualified, ds_qualified, wof_qualified, iw_qualified, cl_qualified,
                            hc_qualified, sh_qualified]:
                season_wipe(region)

            if len(last_stand_tournament['8 Seeds']) < 8:
                enablePrint()
                print(Fore.RED + "NOT ENOUGH 8 SEEDS!" + Fore.RESET)
                write_to_file(error=True, words="NOT ENOUGH 8 SEEDS!")
                time.sleep(1)
            if len(last_stand_tournament['7 Seeds']) < 8:
                enablePrint()
                print(Fore.RED + "NOT ENOUGH 7 SEEDS!" + Fore.RESET)
                write_to_file(error=True, words="NOT ENOUGH 7 SEEDS!")
                time.sleep(1)
            if len(last_stand_tournament['6 Seeds']) < 8:
                enablePrint()
                print(Fore.RED + "NOT ENOUGH 6 SEEDS!" + Fore.RESET)
                write_to_file(error=True, words="NOT ENOUGH 6 SEEDS!")
                time.sleep(1)
            return regional_champs, pre_qualif_tournament, last_stand_tournament
    another = 'y'
    relegated = {} # int keys, team object values
    promoted = {}

    for num in range(SEASONS):


        season_count += 1
        relegated[season_count] = []
        promoted[season_count] = []

        sort_all_players([uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams], manual=False)

        for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
            season_wipe(league)
            for team in league:
                team.history[season_count] = ""
        if season_count != 0:
            season_wipe(uni_teams)
            for team in uni_teams:
                team.history[season_count] = ""

        uni_qualif_g1 = []
        uni_qualif_g2 = []
        regional_champs = []
        last_stand = []
        regional_champs, pqt, last_stand = regional_leagues(dw_teams,sc_teams,ds_teams,wof_teams,iw_teams,cl_teams,hc_teams,sh_teams,season_count=season_count,champ_list=champ_list)

        print(Fore.GREEN + "LAST STAND TOURNAMENT, " + Fore.RESET, end='')
        last_stand_groups = {key: [] for key in range(8)}
        shuffle(last_stand['6 Seeds'])
        shuffle(last_stand['7 Seeds'])
        shuffle(last_stand['8 Seeds'])


        for i in range(8):
            temp7_processed = False
            temp8_processed = False

            other_region = ""
            temp_6 = last_stand['6 Seeds'][i]
            temp_6.accolades['Last-Stand'] += 1
            last_stand_groups[i].append(temp_6)

            for temp_7 in last_stand['7 Seeds']:
                if temp_7.played_region[season_count] != temp_6.played_region[season_count]:
                    other_region = temp_7.played_region[season_count]
                    temp_7.accolades['Last-Stand'] += 1
                    last_stand_groups[i].append(temp_7)
                    last_stand['7 Seeds'].remove(temp_7)
                    temp7_processed = True
                    break

            if not temp7_processed:
                write_to_file(error=True, words=f"temp7 failsafe triggered. Length {len(last_stand['7 Seeds'])}")
                temp_7 = choice(last_stand['7 Seeds'])
                other_region = temp_7.played_region[season_count]
                temp_7.accolades['Last-Stand'] += 1
                last_stand_groups[i].append(temp_7)
                last_stand['7 Seeds'].remove(temp_7)

            for temp_8 in last_stand['8 Seeds']:
                if temp_8.played_region[season_count] not in [temp_6.played_region[season_count], other_region]:
                    temp_8.accolades['Last-Stand'] += 1
                    last_stand_groups[i].append(temp_8)
                    last_stand['8 Seeds'].remove(temp_8)
                    temp8_processed = True
                    break

            if not temp8_processed:
                write_to_file(error=True, words=f"temp8 failsafe triggered. Length {len(last_stand['8 Seeds'])}")
                temp_8 = choice(last_stand['8 Seeds'])
                temp_8.accolades['Last-Stand'] += 1
                last_stand_groups[i].append(temp_8)
                last_stand['8 Seeds'].remove(temp_8)

        for key in last_stand_groups.keys():
            temp_standings = round_robin(last_stand_groups[key], 80, 3)

            temp_standings[0].history[season_count] += f" 1st in Last Stand Group {key} -> Pre-Qualif. Swiss."
            pqt.append(temp_standings[0])

            temp_standings[1].history[season_count] += f" 2nd in Last Stand Group {key} -> Pre-Qualif. Swiss."
            pqt.append(temp_standings[1])

            temp_standings[2].history[season_count] += f" 3rd in Last Stand Group {key} -> Failed to advance."

            season_wipe(temp_standings)

        print(Fore.GREEN + "PRE-QUALIFYING TOURNAMENT, " + Fore.RESET, end='')
        for team in pqt:
            team.accolades['Pre-Qualifying'] += 1
        pre2, pre2_gone = swiss_format(pqt, base_thresh=75, base_margin=8, win_thresh=3, season_count=season_count) # pre-qualifying swiss should go to 75 games with a margin of 8

        for i in range(int(len(pre2)/2)):
            g1 = choice(pre2)
            pre2.remove(g1)
            uni_qualif_g1.append(g1)
            g1.history[season_count] += f" Advanced from PQ Swiss. -> UNI Qualifier."

            g2 = choice(pre2)
            pre2.remove(g2)
            uni_qualif_g2.append(g2)
            g2.history[season_count] += f" Advanced from PQ Swiss. -> UNI Qualifier."
        for team in pre2_gone:
            #1 in 3 chance to get SR pick
            team.second_pick = 3
            team.history[season_count] += f" Failed to advance from PQ Swiss."

        for i in range(8):
            if i % 2 == 0:
                uni_qualif_g1.append(regional_champs[i])
            elif i % 2 == 1:
                uni_qualif_g2.append(regional_champs[i])
        for group in [uni_qualif_g1, uni_qualif_g2]:
            season_wipe(group)

        if len(upl_standings) < 30:
            drop_range1 = [0, 2, 5]
            drop_range2 = [1, 3, 4]
        else:
            drop_range1 = [0, 2, 5, 6, 9, 11, 12,14]
            drop_range2 = [1, 3, 4, 7, 8, 10, 13,15]

        for i in drop_range1:
            uni_qualif_g1.append(upl_standings[i])
            uni_teams.remove(upl_standings[i])
            relegated[season_count].append(upl_standings[i])

        for i in drop_range2:
            uni_qualif_g2.append(upl_standings[i])
            uni_teams.remove(upl_standings[i])
            relegated[season_count].append(upl_standings[i])

        for team in uni_qualif_g1:
            team.accolades['Universal-Qualifying'] += 1
        for team in uni_qualif_g2:
            team.accolades['Universal-Qualifying'] += 1

        print(Fore.GREEN + "GROUP 1 QUALIFYING ROUND, " + Fore.RESET, end='')
        g1_pre_advance = round_robin(uni_qualif_g1, 7, len(uni_qualif_g1), amp=3)
        print(Fore.GREEN + "GROUP 2 QUALIFYING ROUND, " + Fore.RESET, end='')
        g2_pre_advance = round_robin(uni_qualif_g2, 7, len(uni_qualif_g2), amp=3)
        uni_teams.reverse()
        g1_advance = []
        g2_advance = []

        void_draft = []

        for i in range(5): #seeds 1 through 5 who clinch universal league
            g1_advance.append(g1_pre_advance[i])
            uni_teams.append(g1_advance[i])
            g1_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 1 Qualifying. -> UNI League."

            g2_advance.append(g2_pre_advance[i])
            uni_teams.append(g2_advance[i])
            g2_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 2 Qualifying. -> UNI League."
        for i in range(5,14): #6, 7, 8, 9, 10, and 11, 12, 13, and 14 seeds going to play-in tournament
            g1_advance.append(g1_pre_advance[i])
            g1_pre_advance[i].qualifying_group = 1
            g1_pre_advance[i].history[season_count] += f" {ordinal_string(i + 1)} in Group 1 Qualifying. -> Universal Play-In. "

            g2_advance.append(g2_pre_advance[i])
            g2_pre_advance[i].qualifying_group = 2
            g2_pre_advance[i].history[season_count] += f" {ordinal_string(i + 1)} in Group 2 Qualifying. -> Universal Play-In. "
        for i in range(14,len(g1_pre_advance)): #15 seed and below eliminated
            void_draft.append(g1_pre_advance[i])
            void_draft.append(g2_pre_advance[i])

            g1_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 1 Qualifying. Failed to qualify."
            g2_pre_advance[i].history[season_count] += f" {ordinal_string(i+1)} in Group 2 Qualifying. Failed to qualify."

            g1_pre_advance[i].second_pick = 2
            g2_pre_advance[i].second_pick = 2


        for i in range(5):
            if g1_advance[i] in relegated[season_count]:
                relegated[season_count].remove(g1_advance[i])
            else:
                promoted[season_count].append(g1_advance[i])
            if g2_advance[i] in relegated[season_count]:
                relegated[season_count].remove(g2_advance[i])
            else:
                promoted[season_count].append(g2_advance[i])
        #UNIVERSAL PLAY IN
        advanced_play_in = []
        elim_play_in = []


        #IN-GROUP CHAIN, GROUP 1
        print(Fore.BLUE + "(Game 1a) Group 1 no.14 vs Group 1 no.13, loser eliminated" + Fore.RESET)
        igc1_g1W, igc1_g1L = best_of(g1_advance[12], g1_advance[13],
                                   thresh=130, win_by=30,
                                   both_return=True, upset_list=upset_list, upset_count=upset_count,
                                   context=f"S{season_count} Universal Play-In Group 1 Chain Game 1")
        elim_play_in.append(igc1_g1L)

        print(Fore.BLUE + "(Game 2a) Game 1a Winner vs Group 1 no.12, loser eliminated" + Fore.RESET)
        igc1_g2W, igc1_g2L = best_of(igc1_g1W, g1_advance[11],
                                     thresh=130, win_by=30,
                                     both_return=True, upset_list=upset_list, upset_count=upset_count,
                                     context=f"S{season_count} Universal Play-In Group 1 Chain Game 2")
        elim_play_in.append(igc1_g2L)

        print(Fore.BLUE + "(Game 3a) Game 2a Winner vs Group 1 no.11, loser eliminated, winner advances to play-in proper" + Fore.RESET)
        igc1_g3W, igc1_g3L = best_of(igc1_g2W, g1_advance[10],
                                     thresh=130, win_by=30,
                                     both_return=True, upset_list=upset_list, upset_count=upset_count,
                                     context=f"S{season_count} Universal Play-In Group 1 Chain Game 3")
        print(Fore.GREEN + f"{igc1_g3W.name}({igc1_g3W.seed}) represents Group 1 no. 11" + Fore.RESET)
        elim_play_in.append(igc1_g3L)
        #igc1_g3W is the Group 1 11 Seed in the Play-In

        # IN-GROUP CHAIN, GROUP 2
        print(
            Fore.BLUE + "(Game 1b) Group 2 no.14 vs Group 2 no.13, loser eliminated" + Fore.RESET)
        igc2_g1W, igc2_g1L = best_of(g2_advance[12], g2_advance[13],
                                     thresh=130, win_by=30,
                                     both_return=True, upset_list=upset_list, upset_count=upset_count,
                                     context=f"S{season_count} Universal Play-In Group 2 Chain Game 1")
        elim_play_in.append(igc2_g1L)

        print(
            Fore.BLUE + "(Game 2b) Game 1b Winner vs Group 2 no.12, loser eliminated" + Fore.RESET)
        igc2_g2W, igc2_g2L = best_of(igc2_g1W, g2_advance[11],
                                     thresh=130, win_by=30,
                                     both_return=True, upset_list=upset_list, upset_count=upset_count,
                                     context=f"S{season_count} Universal Play-In Group 2 Chain Game 2")
        elim_play_in.append(igc2_g2L)

        print(
            Fore.BLUE + "(Game 3b) Game 2b Winner vs Group 1 no.11, loser eliminated, winner advances to play-in proper" + Fore.RESET)
        igc2_g3W, igc2_g3L = best_of(igc2_g2W, g2_advance[10],
                                     thresh=130, win_by=30,
                                     both_return=True, upset_list=upset_list, upset_count=upset_count,
                                     context=f"S{season_count} Universal Play-In Group 2 Chain Game 3")
        print(Fore.GREEN + f"{igc2_g3W.name}({igc2_g3W.seed}) represents Group 2 no. 11" + Fore.RESET)
        elim_play_in.append(igc2_g3L)
        # igc2_g3W is the Group 2 11 Seed in the Play-In

        # group 1 no.6 vs group 2 no.7
        print(Fore.BLUE + "(Game 1) Group 1 no.6 vs Group 2 no.7, winners advance" + Fore.RESET)
        upi_g1W, upi_g1L = best_of(g1_advance[5], g2_advance[6],
                                   thresh=350, win_by=20,
                                   both_return=True, upset_list=upset_list, upset_count=upset_count,
                                   context=f"S{season_count} Universal Play-In 7v6")
        print(Fore.GREEN + f"{upi_g1W.name} advance" + Fore.RESET)
        advanced_play_in.append(upi_g1W)

        # group 2 no.6 vs group 1 no.7
        print(Fore.BLUE + "(Game 2) Group 2 no.6 vs Group 1 no.7, winners advance" + Fore.RESET)
        upi_g2W, upi_g2L = best_of(g2_advance[5], g1_advance[6],
                                   thresh=350, win_by=20,
                                   both_return=True, upset_list=upset_list, upset_count=upset_count,
                                   context=f"S{season_count} Universal Play-In 7v6")
        print(Fore.GREEN + f"{upi_g2W.name} advance" + Fore.RESET)
        advanced_play_in.append(upi_g2W)


        # group 1 no.10 vs group 2 no.11
        print(Fore.BLUE + "(Game 3) Group 1 no.10 vs Group 2 no.11, losers are eliminated" + Fore.RESET) #g2 no.11 comes from Chain
        upi_g3W, upi_g3L = best_of(g1_advance[9], igc2_g3W,
                                   thresh=350, win_by=20,
                                   both_return=True, upset_list=upset_list, upset_count=upset_count,
                                   context=f"S{season_count} Universal Play-In 11v10")
        print(Fore.GREEN + f"{upi_g3L.name} eliminated" + Fore.RESET)
        elim_play_in.append(upi_g3L)

        # group 2 no.10 vs group 1 no.11
        print(Fore.BLUE + "(Game 4) Group 2 no.10 vs Group 1 no.11, losers are eliminated" + Fore.RESET) #g1 no.11 comes from Chain
        upi_g4W, upi_g4L = best_of(g2_advance[9], igc1_g3W,
                                   thresh=350, win_by=20,
                                   both_return=True, upset_list=upset_list, upset_count=upset_count,
                                   context=f"S{season_count} Universal Play-In 11v10")
        print(Fore.GREEN + f"{upi_g4L.name} eliminated" + Fore.RESET)
        elim_play_in.append(upi_g4L)

        # group 1 no.8 vs group 2 no.9
        print(Fore.BLUE + "(Game 5) Group 1 no.8 vs Group 2 no.9" + Fore.RESET)
        upi_g5W, upi_g5L = best_of(g1_advance[7], g2_advance[8],
                                   thresh=350, win_by=20,
                                   both_return=True, upset_list=upset_list, upset_count=upset_count,
                                   context=f"S{season_count} Universal Play-In 9v8")
        # group 2 no.8 vs group 1 no.9
        print(Fore.BLUE + "(Game 6) Group 2 no.8 vs Group 1 no.9" + Fore.RESET)
        upi_g6W, upi_g6L = best_of(g2_advance[7], g1_advance[8],
                                   thresh=350, win_by=20,
                                   both_return=True, upset_list=upset_list, upset_count=upset_count,
                                   context=f"S{season_count} Universal Play-In 9v8")

        # Game 5/6 winners vs Game 1/2 losers (Games 7 and 8)
        if upi_g5W.qualifying_group != upi_g1L.qualifying_group:
            #Game 5 winner vs Game 1 Loser
            print(Fore.BLUE + "(Game 7) Game 5 winner vs Game 1 loser, winners advance" + Fore.RESET)
            upi_g7W, upi_g7L = best_of(upi_g5W, upi_g1L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g7W.name} advance" + Fore.RESET)
            advanced_play_in.append(upi_g7W)
            #Game 6 winner vs Game 2 Loser
            print(Fore.BLUE + "(Game 8) Game 6 winner vs Game 2 loser, winners advance." + Fore.RESET)
            upi_g8W, upi_g8L = best_of(upi_g6W, upi_g2L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g8W.name} advance" + Fore.RESET)
            advanced_play_in.append(upi_g8W)
        else:
            #Game 6 winner vs Game 1 loser
            print(Fore.BLUE + "(Game 7) Game 6 winner vs Game 1 loser, winners advance" + Fore.RESET)
            upi_g7W, upi_g7L = best_of(upi_g6W, upi_g1L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g7W.name} advance" + Fore.RESET)
            advanced_play_in.append(upi_g7W)
            #Game 5 winner vs Game 2 loser
            print(Fore.BLUE + "(Game 8) Game 5 winner vs Game 2 loser, winners advance" + Fore.RESET)
            upi_g8W, upi_g8L = best_of(upi_g5W, upi_g2L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g8W.name} advance" + Fore.RESET)
            advanced_play_in.append(upi_g8W)

        #Game 3/4 winners vs Game 5/6 losers (Games 9 and 10)
        if upi_g3W.qualifying_group != upi_g5L.qualifying_group:
            #Game 3 winner vs Game 5 Loser
            print(Fore.BLUE + "(Game 9) Game 3 winner vs Game 5 loser, loser is eliminated" + Fore.RESET)
            upi_g9W, upi_g9L = best_of(upi_g3W, upi_g5L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g9L.name} eliminated" + Fore.RESET)
            elim_play_in.append(upi_g9L)
            #Game 4 winner vs Game 6 Loser
            print(Fore.BLUE + "(Game 10) Game 4 winner vs Game 6 loser, loser is eliminated" + Fore.RESET)
            upi_g10W, upi_g10L = best_of(upi_g4W, upi_g6L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g10L.name} eliminated" + Fore.RESET)
            elim_play_in.append(upi_g10L)
        else:
            # Game 4 winner vs Game 5 Loser
            print(Fore.BLUE + "(Game 9) Game 4 winner vs Game 5 loser, loser is eliminated" + Fore.RESET)
            upi_g9W, upi_g9L = best_of(upi_g4W, upi_g5L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g9L.name} eliminated" + Fore.RESET)
            elim_play_in.append(upi_g9L)

            # Game 3 winner vs Game 6 Loser
            print(Fore.BLUE + "(Game 10) Game 3 winner vs Game 6 loser, loser is eliminated" + Fore.RESET)
            upi_g10W, upi_g10L = best_of(upi_g3W, upi_g6L,
                                         thresh=350, win_by=20,
                                         both_return=True, upset_list=upset_list, upset_count=upset_count,
                                         context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g10L.name} eliminated" + Fore.RESET)
            elim_play_in.append(upi_g10L)

        #Game 7/8 losers vs Game 9/10 winners (Games 11 and 12)
        if upi_g7L.qualifying_group != upi_g9W.qualifying_group:
            #Game 9 winner vs Game 7 loser
            print(Fore.BLUE + "(Game 11) Game 9 winner vs Game 7 loser, winner advances/loser is eliminated" + Fore.RESET)
            upi_g11W, upi_g11L = best_of(upi_g9W, upi_g7L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g11W.name} advance\n{upi_g11L.name} eliminated" + Fore.RESET)
            advanced_play_in.append(upi_g11W)
            elim_play_in.append(upi_g11L)
            #Game 10 winner vs Game 8 Loser
            print(Fore.BLUE + "(Game 12) Game 10 winner vs Game 8 loser, winner advances/loser is eliminated" + Fore.RESET)
            upi_g12W, upi_g12L = best_of(upi_g10W, upi_g8L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g12W.name} advance\n{upi_g12L.name} eliminated" + Fore.RESET)
            advanced_play_in.append(upi_g12W)
            elim_play_in.append(upi_g12L)
            print(Fore.RED + "END OF UNIVERSAL PLAY-IN" + Fore.RESET)
        else:
            # Game 10 winner vs Game 7 loser
            print(Fore.BLUE + "(Game 11) Game 10 winner vs Game 7 loser, winner advances/losers are eliminated" + Fore.RESET)
            upi_g11W, upi_g11L = best_of(upi_g10W, upi_g7L,
                                       thresh=350, win_by=20,
                                       both_return=True, upset_list=upset_list, upset_count=upset_count,
                                       context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g11W.name} advance\n{upi_g11L.name} eliminated" + Fore.RESET)
            advanced_play_in.append(upi_g11W)
            elim_play_in.append(upi_g11L)

            # Game 9 winner vs Game 8 loser
            print(Fore.BLUE + "(Game 12) Game 9 winner vs Game 8 loser, winner advances/losers are eliminated" + Fore.RESET)
            upi_g12W, upi_g12L = best_of(upi_g9W, upi_g8L,
                                         thresh=350, win_by=20,
                                         both_return=True, upset_list=upset_list, upset_count=upset_count,
                                         context=f"S{season_count} Universal Play-In")
            print(Fore.GREEN + f"{upi_g12W.name} advance\n{upi_g12L.name} eliminated" + Fore.RESET)
            advanced_play_in.append(upi_g12W)
            elim_play_in.append(upi_g12L)
            print(Fore.RED + "END OF UNIVERSAL PLAY-IN" + Fore.RESET)

        for temp in [5,6,7,8,9,10]:
            if g1_advance[temp] not in elim_play_in:
                advanced_play_in.append(g1_advance[temp])
            if g2_advance[temp] not in elim_play_in:
                advanced_play_in.append(g2_advance[temp])

        advanced_play_in = list(set(advanced_play_in))

        time.sleep(3)


        for team in advanced_play_in:
            uni_teams.append(team)
            if team in relegated[season_count]:
                relegated[season_count].remove(team)
            else:
                promoted[season_count].append(team)
            team.history[season_count] += "Advanced from Universal Play-In -> UNI League. "
        for team in elim_play_in:
            team.history[season_count] += "Failed to advance from Universal Play-In. "

        season_wipe(uni_teams)
        uni_teams = list(set(uni_teams))
        upl_standings = league_season(uni_teams, use_saved=False, season_count=season_count, stats_list=uni_stats_list,upset_list=upset_list,upset_count=upset_count, region="Universal")

        for team in promoted[season_count]:
            print(f"{team.name} PROMOTED.")
            team.history[season_count] += "\n\tPROMOTED to Universal League."
        for team in relegated[season_count]:
            print(f"{team.name} RELEGATED.")
            team.history[season_count] += "\n\tRELEGATED to "

        clear_file('history')
        clear_file('playerstats')
        clear_file('best_stats')

        reverse_upl_standings = list(reversed(upl_standings))



        #to make MVP specific to regional regular season stats, DO NOT run the season_stats function on each region
        #instead, the stats will be generated in the league_season function following the regular season,
        #and the MVP can be determined by running those stats through the region_mvp function


        get_upsets(upset_list, upset_count)

        second_draft_g1 = []
        second_draft_g2 = []
        second_draft_g3 = []

        third_draft_full = []

        #todo currently, pre-qualifying teams are being given second round picks in a way other than being added to g3
        #every srg3 team is getting an extra second round draft pick, meaning it's being granted to them after failing
        #to qualify for the universal qualifying round
        for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
            for team in league:
                if team not in uni_teams:
                    all_teams.append(team)
                if team.second_pick == 1:
                    second_draft_g1.append(team)
                elif team.second_pick == 2:
                    coin = choice([True, True, False])
                    if coin:
                        team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: SUCCESS"
                        second_draft_g2.append(team)
                    else:
                        team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: FAILURE"
                elif team.second_pick == 3:
                    coin = choice([True, False, False])
                    if coin:
                        team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: SUCCESS"
                        second_draft_g3.append(team)
                    else:
                        team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: FAILURE"
                team.second_pick = 0
                if team.third_pick == 1:
                    third_draft_full.append(team)
                    team.third_pick = 0

        for team in uni_teams:
            all_teams.append(team)
            if team.second_pick == 1 and team not in second_draft_g1:
                second_draft_g1.append(team)
            elif team.second_pick == 2 and team not in second_draft_g2:
                coin = choice([True, True, False])
                if coin:
                    team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: SUCCESS"
                    second_draft_g2.append(team)
                else:
                    team.history[season_count] += "\n\t2 in 3 chance for Second Round pick: FAILURE"
            elif team.second_pick == 3 and team not in second_draft_g3:
                coin = choice([True, False, False])
                if coin:
                    team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: SUCCESS"
                    second_draft_g3.append(team)
                else:
                    team.history[season_count] += "\n\t1 in 3 chance for Second Round pick: FAILURE"
            if team.third_pick == 1 and team not in third_draft_full:
                third_draft_full.append(team)
                team.third_pick = 0


        shuffle(second_draft_g1)
        shuffle(second_draft_g2)
        shuffle(second_draft_g3)
        second_draft_full = second_draft_g1 + second_draft_g2 + second_draft_g3

        shuffle(third_draft_full)

        for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
            for team in league:
                for player in team.players:
                    player.age += 1

        #season_stats_list[season_count] = season_stats(all_teams, season_count, season_stats_list)
        #best_of_stats(season_stats_list, season_count, avg_stats_df=avg_stats_df)
        sort_champions_by_seed()

        #REMINDER: each _stats_list is a DICTIONARY with integer keys for the season number, and the values are a LIST OF PLAYER SEASONS


        if True:
            all_players = []

            for league in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
                league.reverse()
            for team in uni_teams:
                if team in dw_teams:
                    dw_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        dw_teams.append(add)
                        add.history[season_count] += "Darkwing Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Darkwing Regional."
                        dw_teams.insert(0, add)
                elif team in sc_teams:
                    sc_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        sc_teams.append(add)
                        add.history[season_count] += "Shining-Core Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Shining-Core Regional."
                        sc_teams.insert(0, add)
                elif team in ds_teams:
                    ds_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        ds_teams.append(add)
                        add.history[season_count] += "Diamond-Sea Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Diamond-Sea Regional."
                        ds_teams.insert(0, add)
                elif team in wof_teams:
                    wof_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        wof_teams.append(add)
                        add.history[season_count] += "Web-of-Nations Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Web-of-Nations Regional."
                        wof_teams.insert(0, add)
                elif team in iw_teams:
                    iw_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        iw_teams.append(add)
                        add.history[season_count] += "Ice-Wall Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Ice-Wall Regional."
                        iw_teams.insert(0, add)
                elif team in cl_teams:
                    cl_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        cl_teams.append(add)
                        add.history[season_count] += "Candyland Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Candyland Regional."
                        cl_teams.insert(0, add)
                elif team in hc_teams:
                    hc_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        hc_teams.append(add)
                        add.history[season_count] += "Hell's-Circle Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Hell's-Circle Regional."
                        hc_teams.insert(0, add)
                elif team in sh_teams:
                    sh_teams.remove(team)
                    if relegated[season_count]:
                        add = choice(relegated[season_count])
                        relegated[season_count].remove(add)
                        sh_teams.append(add)
                        add.history[season_count] += "Steel-Heart Regional."
                    else:
                        add = Team('Labyrinth')
                        add.history[season_count] = "Introduced into the Steel-Heart Regional."
                        sh_teams.insert(0, add)


            with open("rosters", 'w') as l:
                l.write('')

            write_rosters = False
            if write_rosters:
                with open("rosters", 'a') as w:
                    w.write(f"SEASON NO. {season_count}\n")
                    trans_region = ['Universal', 'Darkwing', 'Shining-Core', 'Diamond-Sea', 'Web-of-Nations', 'Ice-Wall', 'Candyland', "Hell's-Circle", 'Steel-Heart']
                    trans_i = -1
                    for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
                        trans_i +=1
                        for team in league:
                            team.region = trans_region[trans_i]
                            for player in team.players:
                                all_players.append(player)
                    grade_players(all_players,is_team=True)
                    for player in all_players:
                        w.write(f"{str(player)}\n")

            #note: reverse_upl_standings is the correct order from 1st to 36th

            for team in elim_play_in:
                team.second_pick = 1
                void_draft.append(team)
            for team in reverse_upl_standings[20:]:
                void_draft.append(team)
            void_draft.reverse()

            dw_draft = [team for team in dw_teams if team not in void_draft]
            sc_draft = [team for team in sc_teams if team not in void_draft]
            ds_draft = [team for team in ds_teams if team not in void_draft]
            wof_draft = [team for team in wof_teams if team not in void_draft]
            iw_draft = [team for team in iw_teams if team not in void_draft]
            cl_draft = [team for team in cl_teams if team not in void_draft]
            hc_draft = [team for team in hc_teams if team not in void_draft]
            sh_draft = [team for team in sh_teams if team not in void_draft]


            user_draft(dw_draft, season_count, is_regional=True, draft_name="Darkwing Regional draft", league_season_stats=dw_stats_list)
            user_draft(sc_draft, season_count, is_regional=True, draft_name="Shining-Core Regional draft", league_season_stats=sc_stats_list)
            user_draft(ds_draft, season_count, is_regional=True, draft_name="Diamond-Sea Regional draft", league_season_stats=ds_stats_list)
            user_draft(wof_draft, season_count, is_regional=True, draft_name="Web-of-Nations Regional draft", league_season_stats=wof_stats_list)
            user_draft(iw_draft, season_count,  is_regional=True, draft_name="Ice-Wall Regional draft", league_season_stats=iw_stats_list)
            user_draft(cl_draft, season_count, is_regional=True, draft_name="Candyland Regional draft", league_season_stats=cl_stats_list)
            user_draft(hc_draft, season_count,  is_regional=True, draft_name="Hell's-Circle Regional draft", league_season_stats=hc_stats_list)
            user_draft(sh_draft, season_count, is_regional=True, draft_name="Steel-Heart Regional draft", league_season_stats=sh_stats_list)

            user_draft(second_draft_full, season_count, second=True, draft_name="Secondary draft")
            user_draft(third_draft_full, season_count, third=True, draft_name="Tertiary draft")

            upl_draft = reverse_upl_standings[0:20]
            upl_draft.reverse()
            #This takes the top 20 from the UPL (those NOT on the chopping block) and reverses them


            #This takes the 16 teams which were eliminated and puts the draft in order from worst to best by reversing
            #void draft includes all teams which failed to qualify from universal qualifying, as well as all teams on the universal chopping block
            upl_draft = remove_duplicates_ordered(upl_draft)


            user_draft(upl_draft, season_count, draft_name='Universal-Draft', league_season_stats=season_stats_list,write_draft=True)
            user_draft(void_draft, season_count, void=True, draft_name='Void-Draft')

            for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams, sh_teams]:
                player_changes(league, season_count=season_count)

            slasher_count = undead_count = reflector_count = clutch_count = inc_count = pp_count = total_player_count = normal_player_count = 0
            clutch_total_mult = inc_total_pct = pp_total_mult = 0
            for league in [uni_teams, dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams, hc_teams,
                           sh_teams]:
                for team in league:
                    for idk in team.players:
                        total_player_count += 1
                        slasher_count += 1 if idk.trait_tag == '$l' else 0
                        undead_count += 1 if idk.trait_tag == 'U-' else 0
                        reflector_count += 1 if idk.trait_tag == 'R#' else 0
                        clutch_count += 1 if idk.trait_tag == 'C%' else 0
                        inc_count += 1 if idk.trait_tag == 'I*' else 0
                        pp_count += 1 if idk.trait_tag == 'Pp' else 0
                        normal_player_count += 1 if idk.trait_tag == 'None' else 0

                    team.print_team_name(season_count)
                    team.print_accolades()
                    team.print_history(season_count)
                    if team.mine:
                        print("\n" + Fore.RED + f"{team.name}\n" + Fore.BLUE + team.history[season_count] + Fore.RESET)

            with open('player_trait_data','a') as file:
                file.write(f"Season No. {season_count}, Total Players: {total_player_count}\n")
                file.write(f"Slashers: {slasher_count} ({round((100*slasher_count/total_player_count),2)}%)\n")
                file.write(f"Undead: {undead_count} ({round((100*undead_count/total_player_count),2)}%)\n")
                file.write(f"Reflectors: {reflector_count} ({round((100*reflector_count/total_player_count),2)}%)\n")
                file.write(f"Clutch Players: {clutch_count} ({round((100*clutch_count/total_player_count),2)}%)\n")
                file.write(f"Inconsistent Players: {inc_count} ({round((100*inc_count/total_player_count),2)}%)\n")
                file.write(f"Playoff Performers: {pp_count} ({round((100*pp_count/total_player_count),2)}%)\n")
                file.write(f"Normal Players (no traits): {normal_player_count} ({round((100*normal_player_count / total_player_count), 2)}%)\n\n")
            finalize_series_data()
            system = [upl_standings,dw_teams,sc_teams,ds_teams,wof_teams,iw_teams,cl_teams,hc_teams,sh_teams,season_count]
            dump_pkl(system)
    # uni_champions.sort(key=lambda x : (x.wins / x.losses), reverse=True)
    end = time.time()
    print(f"\nTotal Execution Time: {round((end-start)/60)} minutes, {round(((end-start)%60),2)} seconds.")

def test_main():
    import numpy as np



    def generate_list(target_total, min_value, max_value, round_to=0):
        final_list = []

        coefficients = np.array([9, 7, 6, 6, 4, 4])

        num_lists = 1

        valid_list_found = False

        while not valid_list_found:

            if round_to == 0:
                initial_values = np.random.randint(min_value, max_value + 1, size=5)
            else:
                initial_values = np.round(np.random.uniform(min_value, max_value, size=5), round_to)

            partial_sum = np.dot(coefficients[:-1], initial_values)
            remaining_value = (target_total - partial_sum) / coefficients[-1]

            if min_value <= remaining_value <= max_value:
                valid_list_found = True

                final_list = np.append(initial_values, remaining_value)


        return final_list


    model_teams = create_teams(32, 'Test')
    model_teams[0].name = "Test_CLONES"
    for pxyr in model_teams[0].players:
        pxyr.power  = 60
        pxyr.atk_dmg = 45
        pxyr.atk_spd = 9
        pxyr.crit_pct = 0.045
        pxyr.crit_x = 10.7
        pxyr.max_health = 540
        pxyr.spawn_time = 11

    total_power_target = 2160 #round to 0
    total_atk_dmg_target = 1620 #round to 0
    total_atk_spd_target = 324 #round to 0
    total_crit_pct_target = 1.62 #round to 4
    total_crit_x_target = 386 #round to 2
    total_health_target = 19440 #round to 2
    total_spawn_target = 396 #round to 0


    for i in range(1,32): #create model teams 1 through 31
        idx = 0
        gen_power_list = generate_list(total_power_target, 47, 63, 0)
        gen_atk_dmg_list = generate_list(total_atk_dmg_target, 35, 65, 0)
        gen_atk_spd_list = generate_list(total_atk_spd_target, 6, 14, 0)
        gen_crit_pct_list = generate_list(total_crit_pct_target, 0.02, 0.06, 4)
        gen_crit_x_list = generate_list(total_crit_x_target, 6, 12, 2)
        gen_health_list = generate_list(total_health_target, 500, 675, 2)
        gen_spawn_list = generate_list(total_spawn_target, 9, 15, 0)
        for plxr in model_teams[i].players:
            plxr.power = gen_power_list[idx]
            plxr.atk_dmg = gen_atk_dmg_list[idx]
            plxr.atk_spd = gen_atk_spd_list[idx]
            plxr.crit_pct = gen_crit_pct_list[idx]
            plxr.crit_x = gen_crit_x_list[idx]
            plxr.max_health = gen_health_list[idx]
            plxr.spawn_time = gen_spawn_list[idx]
            idx += 1



    all_teams = []

    shuffle(model_teams)

    # Split the list into 6 equal parts
    model_groups = [model_teams[i:i + 5] for i in range(0, len(model_teams), 5)]

    dw_teams = create_teams(100, "Darkwing")
    dw_teams = dw_teams + model_groups[0]
    round_robin(dw_teams, 3, len(dw_teams), is_test=True)

    sc_teams = create_teams(100, "Shining-Core")
    sc_teams = sc_teams + model_groups[1]
    round_robin(sc_teams, 3, len(sc_teams), is_test=True)

    ds_teams = create_teams(100, "Diamond-Sea")
    ds_teams = ds_teams + model_groups[2]
    round_robin(ds_teams, 3, len(ds_teams), is_test=True)

    wof_teams = create_teams(100, "Web-of-Nations")
    wof_teams = wof_teams + model_groups[3]
    round_robin(wof_teams, 3, len(wof_teams), is_test=True)

    iw_teams = create_teams(100, "Ice-Wall")
    iw_teams = iw_teams + model_groups[4]
    round_robin(iw_teams, 3, len(iw_teams), is_test=True)

    cl_teams = create_teams(100, "Candyland")
    cl_teams = cl_teams + model_groups[5]
    round_robin(cl_teams, 3, len(cl_teams), is_test=True)

    for group in [dw_teams, sc_teams, ds_teams, wof_teams, iw_teams, cl_teams]:
        for team in group:
            all_teams.append(team)

    team_season_dataframe(all_teams, 0)



    upset_list = []
    champ_list = []
    clear_file('error_output')
    season_count = 0
    season_stats_list = {}

    #league_season(test_teams, season_count=1, region="Test", upset_count=0, upset_list=upset_list, champ_list=champ_list)
    #season_stats_list[season_count] = season_stats(test_teams, season_count, season_stats_list)
    #region_mvp(season_stats_list, season_count=season_count, region='Test')
    #best_of_stats(season_stats_list, season_count=season_count)

def another_test():
    from stat_functions import clear_all_databases

    clear_all_databases()
    initiate_databases()

    teams = create_teams(10, "Universal", season_count=3)
    round_robin(teams, 1, 10)

def flush_database(db='C:/Users/carte/ControlDataBase.db'):
    with sqlite3.connect(db):
        pass



from stat_functions import clear_all_databases

#clear_all_databases()
initiate_databases()
main()
